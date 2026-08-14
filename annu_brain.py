import sys

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("\n=== Annu AI Brain: zaroori packages missing hain ===")
    print("Ye command chalao:\n  pip install python-dotenv\nUske baad app dobara chalao.\n")
    sys.exit(1)

import os
import re
import logging
import platform
import subprocess
import sqlite3
import time
import webbrowser
import urllib.parse
from datetime import datetime, timedelta
from typing import List
import threading
import schedule

# --------------------------------------------------------------------------
# CORE DEPENDENCY CHECK — runs before anything else touches Tkinter/Gemini/etc.
# On a fresh PC a missing package would otherwise crash with a raw
# ImportError traceback the user can't act on. Instead we collect every
# missing *required* package and print one clear Hinglish install command.
# --------------------------------------------------------------------------
_REQUIRED_MODULES = [
    ("google.genai", "google-genai"),
    ("speech_recognition", "SpeechRecognition"),
    ("pyttsx3", "pyttsx3"),
    ("chromadb", "chromadb"),
    ("duckduckgo_search", "duckduckgo-search"),
    ("openpyxl", "openpyxl"),
    ("PIL", "Pillow"),
    ("tkinter", None),  # stdlib on Windows/Mac; on Linux needs a system package
]

def _check_core_dependencies():
    missing_pip, missing_system = [], []
    for module_name, pip_name in _REQUIRED_MODULES:
        try:
            __import__(module_name)
        except ImportError:
            if pip_name:
                missing_pip.append(pip_name)
            else:
                missing_system.append(module_name)
    if missing_pip or missing_system:
        print("\n=== Annu AI Brain: zaroori packages missing hain ===")
        if missing_pip:
            print("Ye command chalao:\n  pip install " + " ".join(missing_pip))
        if "tkinter" in missing_system:
            print(
                "tkinter missing hai — ye Python ke saath aata hai, lekin Linux par "
                "alag se install karna padta hai:\n"
                "  Ubuntu/Debian: sudo apt install python3-tk\n"
                "  Fedora:        sudo dnf install python3-tkinter"
            )
        print("Sab install karne ke baad app dobara chalao.\n")
        sys.exit(1)

_check_core_dependencies()

from google import genai
import speech_recognition as sr
import pyttsx3
import chromadb
from duckduckgo_search import DDGS
import openpyxl
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import scrolledtext
from PIL import Image, ImageTk

# Optional advanced libs check — each of these unlocks one extra feature
# but the app runs fine without them (with a friendly "install X" message
# shown only if/when that specific feature is actually used).
try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
except ImportError:
    PYAUTOGUI_AVAILABLE = False

try:
    import edge_tts
    import asyncio
    EDGE_TTS_AVAILABLE = True
except ImportError:
    EDGE_TTS_AVAILABLE = False

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except ImportError:
    PYPERCLIP_AVAILABLE = False

try:
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    import pywhatkit
    # Powers both WhatsApp sending and YouTube search-and-play below.
    PYWHATKIT_AVAILABLE = True
except ImportError:
    PYWHATKIT_AVAILABLE = False

try:
    import playsound
    PLAYSOUND_AVAILABLE = True
except ImportError:
    PLAYSOUND_AVAILABLE = False

# ==============================================================================
# ANNU AI BRAIN — ULTIMATE RESEARCH, EXCEL & AUTOMATION EDITION
# ==============================================================================

class StorageManager:
    def __init__(self, base_path="./annu_storage"):
        self.base_path = base_path
        self._init_storage_structure()

    def _init_storage_structure(self):
        subfolders = [
            "memory", "knowledge", "documents", "images", 
            "audio", "video", "embeddings", "models", 
            "plugins", "agents", "cache", "backups", "logs", "reports"
        ]
        for folder in subfolders:
            os.makedirs(os.path.join(self.base_path, folder), exist_ok=True)

class ModelManager:
    """
    Uses the current official 'google-genai' SDK (the old 'google-
    generativeai' package is fully deprecated and its models are being
    shut down — that's exactly what caused the 404 error you hit).

    Model is pinned to the 'gemini-flash-latest' ALIAS rather than a
    dated model name like 'gemini-2.5-flash'. Aliases always point at
    whichever stable Flash model Google currently recommends, so this
    won't 404 again the next time a specific dated model is retired.
    """
    DEFAULT_MODEL = "gemini-flash-latest"

    def __init__(self, api_keys: List[str]):
        self.logger = logging.getLogger("AnnuAI.ModelManager")
        self.api_keys = [k.strip() for k in api_keys if k and k.strip()]
        self.current_key_index = 0
        self._lock = threading.Lock()
        
        if not self.api_keys:
            raise ValueError("At least one valid Gemini API Key is required.")
        
        self._configure_current_key()

    def _configure_current_key(self):
        active_key = self.api_keys[self.current_key_index]
        self.client = genai.Client(api_key=active_key)
        self.logger.info(f"Configured Gemini with API Key index {self.current_key_index + 1}")

    def rotate_key(self) -> bool:
        with self._lock:
            if len(self.api_keys) > 1:
                self.current_key_index = (self.current_key_index + 1) % len(self.api_keys)
                self._configure_current_key()
                self.logger.warning(f"Quota exhausted! Rotated to API Key index {self.current_key_index + 1}")
                return True
            return False

    def get_client(self):
        return self.client

    def new_chat(self):
        return self.client.chats.create(model=self.DEFAULT_MODEL)

class VoiceManager:
    """
    Speaks using edge_tts (natural-sounding Hindi neural voice) when
    available, and transparently falls back to the offline pyttsx3 engine
    if edge_tts isn't installed or a call to it fails (e.g. no internet).
    """
    EDGE_VOICE = "hi-IN-SwaraNeural"

    def __init__(self, gui_instance=None):
        self.logger = logging.getLogger("AnnuAI.Voice")
        self.gui_instance = gui_instance
        self.recognizer = sr.Recognizer()
        self.recognizer.energy_threshold = 300
        self.recognizer.dynamic_energy_threshold = True
        self._audio_cache_dir = os.path.join(os.path.expanduser("~"), ".annu_cache")
        os.makedirs(self._audio_cache_dir, exist_ok=True)

    def speak(self, text: str):
        self.logger.info(f"Speaking: {text}")
        print(f"\nAnnu: {text}")
        if self.gui_instance:
            self.gui_instance.safe_update_display(text, "Annu")

        def _run_tts():
            if EDGE_TTS_AVAILABLE:
                try:
                    self._speak_edge_tts(text)
                    return
                except Exception as e:
                    self.logger.warning(f"edge_tts failed, falling back to pyttsx3: {e}")
            self._speak_pyttsx3(text)

        threading.Thread(target=_run_tts, daemon=True).start()

    def _speak_edge_tts(self, text: str):
        out_path = os.path.join(self._audio_cache_dir, f"tts_{int(time.time()*1000)}.mp3")

        async def _generate():
            communicate = edge_tts.Communicate(text, self.EDGE_VOICE)
            await communicate.save(out_path)

        asyncio.run(_generate())

        try:
            # 'playsound' talks to the system's audio APIs directly
            # (winmm on Windows). This is the primary method because
            # os.startfile()/ShellExecute (the old approach) depends on a
            # file-type association AND on COM being initialized in the
            # calling thread — since this runs inside a background
            # threading.Thread, that's exactly what produced the
            # "WinError -2147221003: Application not found" failure.
            if PLAYSOUND_AVAILABLE:
                playsound.playsound(out_path)
            elif platform.system() == "Windows":
                os.startfile(out_path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["afplay", out_path], shell=False)
            else:
                subprocess.Popen(["xdg-open", out_path], shell=False)
        finally:
            # Best-effort cleanup of old cached clips (keep last 20).
            try:
                files = sorted(
                    (os.path.join(self._audio_cache_dir, f) for f in os.listdir(self._audio_cache_dir)),
                    key=os.path.getmtime,
                )
                for old_file in files[:-20]:
                    os.remove(old_file)
            except Exception:
                pass

    def _speak_pyttsx3(self, text: str):
        try:
            engine = pyttsx3.init()
            voices = engine.getProperty('voices')
            selected_voice = self._pick_female_voice(voices)
            if selected_voice:
                engine.setProperty('voice', selected_voice)
            engine.setProperty('rate', 145)  
            engine.setProperty('volume', 1.0)
            engine.say(text)
            engine.runAndWait()
        except Exception as e:
            self.logger.error(f"TTS speech error: {e}")

    # Female voice names commonly shipped across Windows (SAPI5), macOS,
    # and common Linux TTS packages — checked first so we get an actual
    # named female voice rather than guessing.
    _FEMALE_HINTS = [
        "female", "zira", "hazel", "susan", "samantha", "victoria", "karen",
        "moira", "tessa", "veena", "raveena", "aditi", "heera", "kalpana",
        "hindi", "india",
    ]
    _MALE_HINTS = [
        "male", "david", "mark", "george", "daniel", "alex", "fred",
        "ravi", "rishi", "guy",
    ]

    def _pick_female_voice(self, voices):
        if not voices:
            return None
        # 1) Name/id matches a known female voice.
        for voice in voices:
            name = (getattr(voice, "name", "") or "").lower()
            vid = (getattr(voice, "id", "") or "").lower()
            if any(h in name or h in vid for h in self._FEMALE_HINTS):
                return voice.id
        # 2) Some pyttsx3 backends expose an explicit gender field.
        for voice in voices:
            gender = str(getattr(voice, "gender", "") or "").lower()
            if "female" in gender:
                return voice.id
        # 3) Last resort: skip anything that's clearly a male-named voice,
        # so we at least avoid defaulting to a male voice by accident.
        for voice in voices:
            name = (getattr(voice, "name", "") or "").lower()
            if not any(h in name for h in self._MALE_HINTS):
                return voice.id
        return voices[0].id

    def listen(self) -> str:
        with sr.Microphone() as source:
            print("\n[Listening... Speak now]")
            if self.gui_instance:
                self.gui_instance.safe_update_display("Listening...", "System")
            
            self.recognizer.adjust_for_ambient_noise(source, duration=1.0)
            try:
                audio = self.recognizer.listen(source, timeout=8, phrase_time_limit=12)
                text = self.recognizer.recognize_google(audio, language="hi-IN")
                return text
            except sr.WaitTimeoutError:
                return ""
            except sr.UnknownValueError:
                return ""
            except sr.RequestError as e:
                self.logger.error(f"Speech service error: {e}")
                return ""

class AdvancedWorkspaceManager:
    """Handles professional file creation, desktop saving, and Excel data management."""
    def __init__(self):
        self.desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.logger = logging.getLogger("AnnuAI.Workspace")

    def save_to_desktop(self, filename: str, content: str, extension: str = "txt") -> str:
        try:
            if not filename.endswith(f".{extension}"):
                filename = f"{filename}.{extension}"
            filepath = os.path.join(self.desktop_path, filename)
            
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(content)
            return f"File successfully saved on Desktop as {filename}."
        except Exception as e:
            self.logger.error(f"Save error: {e}")
            return f"Error saving file: {e}"

    def create_professional_excel(self, filename: str, data_rows: List[List[str]], headers: List[str]) -> str:
        try:
            if not filename.endswith(".xlsx"):
                filename = f"{filename}.xlsx"
            filepath = os.path.join(self.desktop_path, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Annu_Report"

            thin_side = openpyxl.styles.Side(style="thin", color="D9D9D9")
            thin_border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            # Header row: bold white text on a solid blue band, centered,
            # with wrapping so long headers don't force a huge column width.
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[1].height = 22

            # Data rows: light alternating banding + wrapped text + borders,
            # so long research paragraphs stay readable instead of spilling
            # off-screen in one endless line.
            for row_idx, row in enumerate(data_rows, start=2):
                ws.append(row)
                band_color = "F2F6FC" if row_idx % 2 == 0 else "FFFFFF"
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.fill = openpyxl.styles.PatternFill(start_color=band_color, end_color=band_color, fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border

            # Column widths sized to content (capped so a long research
            # paragraph doesn't stretch the sheet absurdly wide).
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                max_len = len(str(headers[col_num - 1]))
                for row in data_rows:
                    if col_num - 1 < len(row):
                        max_len = max(max_len, len(str(row[col_num - 1])))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 80)

            ws.freeze_panes = "A2"
            wb.save(filepath)
            return f"Professional Excel sheet '{filename}' successfully saved on Desktop!"
        except Exception as e:
            self.logger.error(f"Excel generation error: {e}")
            return f"Excel generation error: {e}"

class OnlineKnowledgeManager:
    def __init__(self):
        self.ddgs = DDGS()
        self.logger = logging.getLogger("AnnuAI.Web")

    def search_web(self, query: str) -> dict:
        try:
            results = list(self.ddgs.text(query, max_results=5))
            if not results:
                return {"status": "failed", "context": ""}
            
            context = "--- DEEP RESEARCH & INTERNET DATA ---\n"
            for idx, res in enumerate(results):
                context += f"Source {idx+1}: {res.get('title','')}\nDetails: {res.get('body','')}\n\n"
            return {"status": "success", "context": context}
        except Exception as e:
            self.logger.error(f"Search error: {e}")
            return {"status": "failed", "context": ""}

    def search_structured(self, query: str, max_results: int = 5) -> list:
        """Same DuckDuckGo search as search_web, but returns the raw
        [{'title':..., 'body':...}, ...] list instead of one flattened
        blob — needed when the caller wants to lay results out as
        separate rows (e.g. the morning briefing Excel report)."""
        try:
            return list(self.ddgs.text(query, max_results=max_results)) or []
        except Exception as e:
            self.logger.error(f"Structured search error: {e}")
            return []

class ToolManager:
    """
    SECURITY NOTE:
    User/voice input is never passed to a shell. All subprocess calls use
    argument lists with shell=False, and any free-form 'target' text is
    whitelisted to safe characters before being used anywhere (including
    in filenames or process names) to prevent shell/command injection.
    """

    # Known apps mapped to safe, fixed launch commands (no shell involved).
    # Different OSes launch the same app differently, so pick the right
    # command set once at class-definition time based on the host platform —
    # this is what lets the same script work unmodified on Windows/Mac/Linux.
    if platform.system() == "Windows":
        KNOWN_APPS = {
            "chrome": ["chrome"],
            "notepad": ["notepad.exe"],
            "excel": ["excel.exe"],
            "code": ["code"],
            "vs code": ["code"],
        }
    elif platform.system() == "Darwin":
        KNOWN_APPS = {
            "chrome": ["open", "-a", "Google Chrome"],
            "notepad": ["open", "-a", "TextEdit"],
            "excel": ["open", "-a", "Microsoft Excel"],
            "code": ["open", "-a", "Visual Studio Code"],
            "vs code": ["open", "-a", "Visual Studio Code"],
        }
    else:
        KNOWN_APPS = {
            "chrome": ["google-chrome"],
            "notepad": ["gedit"],
            "excel": ["libreoffice", "--calc"],
            "code": ["code"],
            "vs code": ["code"],
        }

    _SAFE_NAME_RE = re.compile(r"[^a-zA-Z0-9 _\-\.]")

    # Common websites people mean by "kholo" even though they're not
    # installed desktop apps. Checked before KNOWN_APPS/os.startfile so
    # "youtube kholo" actually opens youtube.com instead of failing to
    # find a program literally named "youtube" and leaving Gemini to
    # improvise a generic "I can't open apps" answer.
    WEBSITE_SHORTCUTS = {
        "youtube": "https://www.youtube.com",
        "gmail": "https://mail.google.com",
        "google": "https://www.google.com",
        "facebook": "https://www.facebook.com",
        "instagram": "https://www.instagram.com",
        "whatsapp web": "https://web.whatsapp.com",
        "twitter": "https://twitter.com",
        "amazon": "https://www.amazon.in",
        "flipkart": "https://www.flipkart.com",
        "maps": "https://maps.google.com",
        "netflix": "https://www.netflix.com",
        "linkedin": "https://www.linkedin.com",
    }

    # Shopping sites with real product-search URLs (not just the homepage).
    # Deliberately opens a search results page and stops there — it never
    # adds to cart, logs in, or completes checkout/payment on its own, so a
    # misheard voice command can't accidentally spend real money.
    SHOPPING_SITES = {
        "amazon": "https://www.amazon.in/s?k={q}",
        "flipkart": "https://www.flipkart.com/search?q={q}",
    }

    def shop_search(self, platform: str, query: str) -> str:
        platform = (platform or "").lower().strip()
        query = self._sanitize(query)
        url_template = self.SHOPPING_SITES.get(platform)
        if not url_template:
            return f"'{platform}' abhi supported nahi hai — Amazon ya Flipkart try karo."
        if not query:
            return f"{platform} par kya khojna hai, product ka naam bataiye."
        try:
            url = url_template.format(q=urllib.parse.quote(query))
            webbrowser.open(url)
            return (
                f"{platform} par '{query}' search kar diya hai. Order/payment tumhe khud "
                f"confirm karni hogi — main checkout ya payment apne aap nahi karti."
            )
        except Exception as e:
            self.logger.error(f"Shop search error: {e}")
            return "Shopping search karne mein error aa gayi."

    def __init__(self, model_manager):
        self.logger = logging.getLogger("AnnuAI.ToolManager")
        self.workspace = AdvancedWorkspaceManager()
        self.vision_manager = ScreenVisionManager(model_manager) if PYAUTOGUI_AVAILABLE else None

    @classmethod
    def _sanitize(cls, text: str) -> str:
        """Strip anything that isn't alphanumeric/space/dash/underscore/dot."""
        cleaned = cls._SAFE_NAME_RE.sub("", text or "").strip()
        return cleaned

    def execute_app_command(self, action_type: str, target: str) -> str:
        action_type = (action_type or "").lower().strip()
        raw_target = (target or "").lower().strip()
        safe_target = self._sanitize(raw_target)

        if not safe_target and action_type != "play":
            return "Kis app ka naam bataiye, samajh nahi paayi."

        if action_type == "open":
            try:
                if "spotify" in safe_target:
                    # Use the OS opener instead of a shell command string.
                    if platform.system() == "Windows":
                        os.startfile("spotify:")
                    else:
                        webbrowser.open("https://open.spotify.com")
                    return "Spotify khol di gayi hai."

                for site_key, url in self.WEBSITE_SHORTCUTS.items():
                    if site_key in safe_target:
                        webbrowser.open(url)
                        return f"{site_key} khol diya gaya hai (browser mein)."

                for key, cmd in self.KNOWN_APPS.items():
                    if key in safe_target:
                        subprocess.Popen(cmd, shell=False)
                        return f"{key} khol diya gaya hai."

                # Unknown/free-form app name: try the OS-native opener only,
                # never a shell string built from user text.
                if platform.system() == "Windows":
                    os.startfile(safe_target)
                else:
                    subprocess.Popen([safe_target], shell=False)
                return f"{safe_target} open kar diya gaya hai."
            except FileNotFoundError:
                return f"'{safe_target}' naam ka app nahi mila."
            except Exception as e:
                self.logger.error(f"App open error: {e}")
                return "App open karne mein error aa gayi."

        elif action_type == "play":
            try:
                if not safe_target or safe_target in ["music", "gana", "song"]:
                    safe_target = "latest trending hindi songs"
                # Proper URL-encoding instead of a naive space replace,
                # and opened via the standard browser API (no shell).
                query = urllib.parse.quote(safe_target)
                url = f"https://open.spotify.com/search/{query}"
                webbrowser.open(url)
                return f"Spotify par '{safe_target}' search kiya ja raha hai."
            except Exception as e:
                self.logger.error(f"Play error: {e}")
                return "Play karne mein error aa gayi."

        elif action_type == "play_youtube":
            if not PYWHATKIT_AVAILABLE:
                return "YouTube par direct play karne ke liye 'pywhatkit' install karo: pip install pywhatkit"
            try:
                if not safe_target or safe_target in ["music", "gana", "song"]:
                    safe_target = "latest trending hindi songs"
                # playonyt opens the browser AND auto-plays the first
                # matching result — unlike a plain search link, no extra
                # click needed.
                pywhatkit.playonyt(safe_target)
                return f"YouTube par '{safe_target}' play kar rahi hoon."
            except Exception as e:
                self.logger.error(f"YouTube play error: {e}")
                return "YouTube par play karne mein error aa gayi."

        elif action_type == "close":
            try:
                if not safe_target:
                    return "Kaunsa app band karna hai, naam bataiye."
                process_name = f"{safe_target}.exe"
                if platform.system() == "Windows":
                    # Argument list, no shell=True: 'target' cannot break out
                    # into extra shell commands even if it contains
                    # metacharacters (those were already stripped above too).
                    subprocess.call(["taskkill", "/F", "/IM", process_name], shell=False)
                else:
                    subprocess.call(["pkill", "-f", safe_target], shell=False)
                return f"{safe_target} close kar diya gaya hai."
            except Exception as e:
                self.logger.error(f"Close error: {e}")
                return "App close karne mein error aa gayi."

        return "Command execute nahi ho saki."

class ScreenVisionManager:
    def __init__(self, model_manager):
        self.model_manager = model_manager

    def analyze_screen(self, goal: str) -> str:
        try:
            screenshot = pyautogui.screenshot()  # PIL Image — passed straight in, no upload step needed.
            client = self.model_manager.get_client()
            response = client.models.generate_content(
                model=self.model_manager.DEFAULT_MODEL,
                contents=[screenshot, f"Analyze this screen screenshot. Goal: '{goal}'. Give specific insights."],
            )
            return response.text
        except Exception as e:
            return f"Vision error: {e}"

class MemoryManager:
    def __init__(self, db_path="annu_memory.db"):
        self.db_path = db_path
        self._init_db()

    def _init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS long_term_memory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                category TEXT,
                content TEXT UNIQUE,
                confidence REAL
            )
        ''')
        conn.commit()
        conn.close()

    def process_and_store(self, content: str, category: str = "general"):
        timestamp = datetime.now().isoformat()
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT OR IGNORE INTO long_term_memory (timestamp, category, content, confidence) VALUES (?, ?, ?, ?)",
                (timestamp, category, content, 0.9)
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logging.getLogger("AnnuAI.Memory").error(f"DB store error: {e}")


class SemanticMemoryManager:
    """
    Vector-search memory using chromadb. Lets Annu recall relevant past
    conversations automatically (not just the current session's history),
    so answers stay context-aware across restarts — the core of making
    'any task' easier over time instead of starting from zero each time.
    """
    def __init__(self, persist_path="./annu_storage/embeddings"):
        self.logger = logging.getLogger("AnnuAI.SemanticMemory")
        os.makedirs(persist_path, exist_ok=True)
        try:
            self.client = chromadb.PersistentClient(path=persist_path)
            self.collection = self.client.get_or_create_collection(name="annu_conversations")
            self.available = True
        except Exception as e:
            self.logger.error(f"chromadb init failed, semantic recall disabled: {e}")
            self.available = False

    def add_memory(self, text: str, role: str = "user"):
        if not self.available or not text.strip():
            return
        try:
            doc_id = f"{role}_{int(time.time()*1000)}_{threading.get_ident()}"
            self.collection.add(
                documents=[text],
                metadatas=[{"role": role, "timestamp": datetime.now().isoformat()}],
                ids=[doc_id],
            )
        except Exception as e:
            self.logger.error(f"Semantic memory add error: {e}")

    def recall(self, query: str, top_k: int = 3) -> str:
        if not self.available or not query.strip():
            return ""
        try:
            results = self.collection.query(query_texts=[query], n_results=top_k)
            docs = results.get("documents", [[]])[0]
            if not docs:
                return ""
            return "\n".join(f"- {d}" for d in docs)
        except Exception as e:
            self.logger.error(f"Semantic memory recall error: {e}")
            return ""


class TaskManager:
    """SQLite-backed to-do list and reminders, with due-time voice alerts."""
    def __init__(self, db_path="annu_tasks.db"):
        self.db_path = db_path
        self.logger = logging.getLogger("AnnuAI.Tasks")
        self._init_db()

    def _init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                description TEXT NOT NULL,
                remind_at TEXT,
                created_at TEXT NOT NULL,
                done INTEGER NOT NULL DEFAULT 0,
                alerted INTEGER NOT NULL DEFAULT 0
            )
        ''')
        conn.commit()
        conn.close()

    def add_task(self, description: str, remind_at: str = None) -> int:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO tasks (description, remind_at, created_at) VALUES (?, ?, ?)",
            (description.strip(), remind_at, datetime.now().isoformat()),
        )
        task_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return task_id

    def list_tasks(self, include_done: bool = False):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        if include_done:
            cursor.execute("SELECT id, description, remind_at, done FROM tasks ORDER BY id DESC")
        else:
            cursor.execute("SELECT id, description, remind_at, done FROM tasks WHERE done = 0 ORDER BY id DESC")
        rows = cursor.fetchall()
        conn.close()
        return rows

    def complete_task(self, task_id: int) -> bool:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("UPDATE tasks SET done = 1 WHERE id = ?", (task_id,))
        changed = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return changed

    def delete_task(self, task_id: int) -> bool:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
        changed = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return changed

    def due_tasks(self):
        """Tasks whose remind_at has passed and haven't been alerted yet."""
        now_iso = datetime.now().isoformat()
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, description FROM tasks WHERE done = 0 AND alerted = 0 "
            "AND remind_at IS NOT NULL AND remind_at <= ?",
            (now_iso,),
        )
        rows = cursor.fetchall()
        if rows:
            ids = [r[0] for r in rows]
            cursor.executemany("UPDATE tasks SET alerted = 1 WHERE id = ?", [(i,) for i in ids])
            conn.commit()
        conn.close()
        return rows


class FileManager:
    """
    Searches for and opens files, restricted to the user's own Desktop,
    Documents and Downloads folders. Every path is resolved to an absolute
    real path and checked against that whitelist before any file
    operation — this blocks path-traversal (e.g. '..' tricks) from ever
    reaching a file outside those folders.
    """
    def __init__(self):
        home = os.path.expanduser("~")
        self.allowed_dirs = [
            os.path.realpath(os.path.join(home, "Desktop")),
            os.path.realpath(os.path.join(home, "Documents")),
            os.path.realpath(os.path.join(home, "Downloads")),
        ]
        self.logger = logging.getLogger("AnnuAI.FileManager")

    def _is_within_allowed_dirs(self, path: str) -> bool:
        real = os.path.realpath(path)
        return any(
            real == base or real.startswith(base + os.sep)
            for base in self.allowed_dirs
        )

    def search(self, name_fragment: str, max_results: int = 8):
        name_fragment = re.sub(r"[^a-zA-Z0-9 _\-\.]", "", name_fragment or "").strip().lower()
        if not name_fragment:
            return []
        matches = []
        for base in self.allowed_dirs:
            if not os.path.isdir(base):
                continue
            for dirpath, _dirnames, filenames in os.walk(base):
                for fname in filenames:
                    if name_fragment in fname.lower():
                        matches.append(os.path.join(dirpath, fname))
                        if len(matches) >= max_results:
                            return matches
        return matches

    def open_file(self, path: str) -> str:
        if not self._is_within_allowed_dirs(path):
            self.logger.warning(f"Blocked attempt to open path outside allowed dirs: {path}")
            return "Ye file allowed folders (Desktop/Documents/Downloads) ke bahar hai, nahi khol sakti."
        if not os.path.isfile(path):
            return "Ye file maujood nahi hai."
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", path], shell=False)
            else:
                subprocess.Popen(["xdg-open", path], shell=False)
            return f"File khol di: {os.path.basename(path)}"
        except Exception as e:
            self.logger.error(f"File open error: {e}")
            return "File open karne mein error aa gayi."


class SystemInfoManager:
    """Battery, CPU/RAM, date and time — degrades gracefully without psutil."""
    def status(self) -> str:
        parts = [f"Aaj ki date: {datetime.now().strftime('%d-%m-%Y')}, time: {datetime.now().strftime('%I:%M %p')}"]
        if PSUTIL_AVAILABLE:
            try:
                parts.append(f"CPU usage: {psutil.cpu_percent(interval=0.5)}%")
                parts.append(f"RAM usage: {psutil.virtual_memory().percent}%")
                battery = psutil.sensors_battery()
                if battery:
                    plugged = "charging" if battery.power_plugged else "on battery"
                    parts.append(f"Battery: {battery.percent}% ({plugged})")
            except Exception:
                pass
        else:
            parts.append("(CPU/RAM/battery details ke liye 'psutil' library install karo: pip install psutil)")
        return "\n".join(parts)


class ClipboardManager:
    def copy(self, text: str) -> str:
        if not PYPERCLIP_AVAILABLE:
            return "Clipboard feature ke liye 'pyperclip' install karo: pip install pyperclip"
        try:
            pyperclip.copy(text)
            return "Clipboard mein copy kar diya."
        except Exception as e:
            return f"Clipboard error: {e}"

    def read(self) -> str:
        if not PYPERCLIP_AVAILABLE:
            return "Clipboard feature ke liye 'pyperclip' install karo: pip install pyperclip"
        try:
            return pyperclip.paste() or "Clipboard khaali hai."
        except Exception as e:
            return f"Clipboard error: {e}"


class WhatsAppManager:
    """
    Sends WhatsApp messages by driving WhatsApp Web in your normal browser
    session (via pywhatkit) — the same thing you'd do by hand, just
    automated. No unofficial API keys, no access to anyone else's data.

    Responsible-use note: this opens a real WhatsApp Web tab and sends to
    a number you provide explicitly each time. It's built for one-off
    personal messages ("tell Priya I'm running late"), not bulk/automated
    broadcasting — sending unsolicited bulk messages violates WhatsApp's
    terms, so this class deliberately has no loop/batch-send method.
    """
    def __init__(self):
        self.logger = logging.getLogger("AnnuAI.WhatsApp")

    @staticmethod
    def _normalize_number(raw: str) -> str:
        digits = re.sub(r"[^\d+]", "", raw or "")
        if not digits.startswith("+"):
            # Rohit is India-based, so a bare 10-digit number is assumed to
            # be a local Indian number; anything else just gets a '+'.
            digits = "+91" + digits if len(digits) == 10 else "+" + digits
        return digits

    def send_message(self, raw_number: str, message: str) -> str:
        if not PYWHATKIT_AVAILABLE:
            return "WhatsApp feature ke liye 'pywhatkit' install karo: pip install pywhatkit"
        if not raw_number or not (message or "").strip():
            return "Number aur message dono chahiye — jaise: 'whatsapp 9876543210 kal milte hain'"
        try:
            number = self._normalize_number(raw_number)
            pywhatkit.sendwhatmsg_instantly(number, message.strip(), wait_time=15, tab_close=True)
            return f"WhatsApp message {number} par bhej diya."
        except Exception as e:
            self.logger.error(f"WhatsApp send error: {e}")
            return f"WhatsApp message bhejne mein error aa gayi: {e}. (Browser mein WhatsApp Web login hona chahiye.)"


class OCRManager:
    """
    Extracts text from images (screenshots, scanned forms, photographed
    documents) — built for data-entry workflows: point it at an invoice/
    form image and get back editable text instead of retyping it by hand.
    Requires the 'pytesseract' package AND the Tesseract-OCR binary
    installed on the system (pytesseract is just a wrapper around it).
    """
    def __init__(self):
        self.logger = logging.getLogger("AnnuAI.OCR")

    def extract_text(self, image_path: str) -> str:
        if not OCR_AVAILABLE:
            return ("OCR ke liye 'pytesseract' install karo (pip install pytesseract) "
                    "aur Tesseract-OCR engine bhi system par install hona chahiye.")
        try:
            if not os.path.isfile(image_path):
                return "Image file nahi mili."
            img = Image.open(image_path)
            text = pytesseract.image_to_string(img, lang="eng+hin")
            text = text.strip()
            if not text:
                return "Is image mein koi readable text nahi mila."
            return text
        except Exception as e:
            self.logger.error(f"OCR error: {e}")
            return f"OCR karte waqt error aa gayi: {e}"

class AnnuAIBrainGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Annu AI Brain - Ultimate Research & Automation Companion")
        self.root.geometry("520x720")
        self.root.configure(bg="#1e1e2e")
        self.running = False  # set True in start_background_threads; avoids
                               # a race if a background loop somehow checks
                               # this before that call happens.

        self._setup_logging()
        
        api_keys = []
        for i in range(1, 5):
            k = os.getenv(f"GEMINI_API_KEY_{i}")
            if k:
                api_keys.append(k)
        
        if not api_keys:
            single_key = os.getenv("GEMINI_API_KEY")
            if single_key:
                api_keys.append(single_key)
            else:
                # Use a GUI dialog instead of blocking console input() —
                # input() would hang forever if the app is launched without
                # an attached terminal (e.g. double-clicked .exe/.pyw).
                from tkinter import simpledialog, messagebox
                user_key = simpledialog.askstring(
                    "Gemini API Key required",
                    "Apni Google Gemini API Key yahan enter karein:",
                    show="*",
                    parent=self.root,
                )
                if user_key and user_key.strip():
                    api_keys.append(user_key.strip())
                else:
                    messagebox.showerror("Annu AI Brain", "API Key ke bina app start nahi ho sakti.")
                    raise ValueError("API Key is required.")

        self.model_manager = ModelManager(api_keys)
        self.voice = VoiceManager(gui_instance=self)
        self.workspace = AdvancedWorkspaceManager()
        self.web = OnlineKnowledgeManager()
        self.tools = ToolManager(self.model_manager)
        self.memory = MemoryManager()
        self.semantic_memory = SemanticMemoryManager()
        self.tasks = TaskManager()
        self.files = FileManager()
        self.sysinfo = SystemInfoManager()
        self.clipboard = ClipboardManager()
        self.ocr = OCRManager()
        self.whatsapp = WhatsAppManager()

        self.chat_session = self.model_manager.new_chat()

        self.create_widgets()
        self.root.after(1000, self.start_background_threads)

    def _setup_logging(self):
        logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
        self.logger = logging.getLogger("AnnuAI")

    def create_widgets(self):
        avatar_frame = tk.Frame(self.root, bg="#252538", bd=2, relief=tk.GROOVE)
        avatar_frame.pack(pady=10, padx=10, fill=tk.X)

        try:
            image_path = "annu_avatar.png"
            if not os.path.exists(image_path):
                image_path = "annu_avatar.jpg"
            img = Image.open(image_path)
            img = img.resize((140, 140), Image.Resampling.LANCZOS)
            self.avatar_img = ImageTk.PhotoImage(img)
            self.avatar_label = tk.Label(avatar_frame, image=self.avatar_img, bg="#252538")
            self.avatar_label.pack(side=tk.LEFT, padx=10, pady=10)
        except Exception:
            self.avatar_label = tk.Label(avatar_frame, text="✨ ANNU-CHAN ✨\n[Research Pro]", font=("Arial", 10, "bold"), fg="#ff79c6", bg="#252538")
            self.avatar_label.pack(side=tk.LEFT, padx=10, pady=10)

        status_text = tk.Label(avatar_frame, text="Annu AI Brain v3.6\nStatus: Online & Ready 🚀", font=("Arial", 9, "bold"), fg="#89b4fa", bg="#252538", justify=tk.LEFT)
        status_text.pack(side=tk.LEFT, padx=10, pady=10)

        chat_frame = tk.Frame(self.root, bg="#1e1e2e")
        chat_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        self.chat_display = scrolledtext.ScrolledText(chat_frame, wrap=tk.WORD, bg="#11111b", fg="#cdd6f4", font=("Arial", 10))
        self.chat_display.pack(fill=tk.BOTH, expand=True)
        self.chat_display.config(state=tk.DISABLED)

        input_frame = tk.Frame(self.root, bg="#1e1e2e")
        input_frame.pack(pady=10, padx=10, fill=tk.X)

        self.text_input = tk.Entry(input_frame, font=("Arial", 11), bg="#313244", fg="#cdd6f4", insertbackground="white")
        self.text_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.text_input.bind("<Return>", self.on_text_submit)

        send_btn = tk.Button(input_frame, text="Send", font=("Arial", 10, "bold"), bg="#89b4fa", fg="#11111b", command=self.on_text_submit)
        send_btn.pack(side=tk.RIGHT)

        tools_frame = tk.Frame(self.root, bg="#1e1e2e")
        tools_frame.pack(pady=(0, 10), padx=10, fill=tk.X)

        ocr_btn = tk.Button(
            tools_frame, text="📷 Image se Text (OCR)", font=("Arial", 9, "bold"),
            bg="#a6e3a1", fg="#11111b", command=self.on_ocr_button
        )
        ocr_btn.pack(side=tk.LEFT, padx=(0, 5))

        tasks_btn = tk.Button(
            tools_frame, text="📋 Tasks", font=("Arial", 9, "bold"),
            bg="#f9e2af", fg="#11111b", command=self.on_show_tasks
        )
        tasks_btn.pack(side=tk.LEFT)

    def on_ocr_button(self):
        from tkinter import filedialog
        image_path = filedialog.askopenfilename(
            title="Image chuniye jisse text nikaalna hai",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.bmp *.tiff"), ("All files", "*.*")],
        )
        if not image_path:
            return
        self.safe_update_display(f"[Image selected: {os.path.basename(image_path)}] OCR chal raha hai...", "System")
        threading.Thread(target=self._run_ocr, args=(image_path,), daemon=True).start()

    def _run_ocr(self, image_path: str):
        extracted = self.ocr.extract_text(image_path)
        self.safe_update_display(extracted, "OCR Result")
        try:
            save_result = self.workspace.save_to_desktop(
                f"OCR_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}", extracted, "txt"
            )
            self.safe_update_display(save_result, "System")
        except Exception as e:
            self.logger.error(f"OCR save error: {e}")

    def on_show_tasks(self):
        rows = self.tasks.list_tasks()
        if not rows:
            self.safe_update_display("Koi pending task nahi hai.", "System")
            return
        lines = []
        for task_id, description, remind_at, _done in rows:
            when = f" (yaad: {remind_at})" if remind_at else ""
            lines.append(f"#{task_id}: {description}{when}")
        self.safe_update_display("\n".join(lines), "Tasks")

    def safe_update_display(self, message: str, sender: str = "Annu"):
        self.root.after(0, lambda: self._update_chat_display_direct(message, sender))

    def _update_chat_display_direct(self, message: str, sender: str):
        self.chat_display.config(state=tk.NORMAL)
        self.chat_display.insert(tk.END, f"{sender}: {message}\n")
        self.chat_display.see(tk.END)
        self.chat_display.config(state=tk.DISABLED)

    def on_text_submit(self, event=None):
        text = self.text_input.get().strip()
        if text:
            self.text_input.delete(0, tk.END)
            self.safe_update_display(text, "You")
            threading.Thread(target=self.handle_user_input, args=(text,), daemon=True).start()

    def handle_user_input(self, user_input: str):
        lower = user_input.lower()
        if any(word in lower for word in ['exit', 'quit', 'shut down app']):
            self.voice.speak("Shutting down Annu AI Brain. Goodbye!")
            self.root.quit()
            return

        tool_output = ""
        web_context = ""

        # Reminders/tasks: "5 minute baad yaad dilana ...", "task add karo ..."
        reminder_match = re.search(
            r"(\d+)\s*(minute|minutes|min|ghante|ghanta|hour|hours)\s*(baad|later|mein)?.*?"
            r"(?:yaad dilana|remind me|reminder)\s*(?:ki|that|:)?\s*(.*)",
            lower,
        )
        if reminder_match:
            amount = int(reminder_match.group(1))
            unit = reminder_match.group(2)
            description = reminder_match.group(4).strip() or user_input
            minutes = amount * 60 if unit.startswith(("hour", "ghant")) else amount
            remind_at = (datetime.now() + timedelta(minutes=minutes)).isoformat()
            self.tasks.add_task(description, remind_at)
            tool_output = f"Theek hai, {amount} {unit} baad yaad dila dungi: '{description}'"
        elif "yaad dilana" in lower or "remind me" in lower or "reminder" in lower:
            self.tasks.add_task(user_input, remind_at=None)
            tool_output = "Task list mein add kar diya (koi fixed time set nahi kiya gaya)."
        elif "tasks dikhao" in lower or "task list" in lower or "mere task" in lower:
            rows = self.tasks.list_tasks()
            tool_output = "\n".join(f"#{i}: {d}" for i, d, _r, _done in rows) or "Koi pending task nahi hai."

        # File search / open: "file dhundo invoice", "file kholo report.pdf"
        elif "file dhundo" in lower or "find file" in lower or "search file" in lower:
            fragment = lower.replace("file dhundo", "").replace("find file", "").replace("search file", "").strip()
            matches = self.files.search(fragment)
            tool_output = "\n".join(matches) if matches else "Koi matching file nahi mili (Desktop/Documents/Downloads mein dekha)."
        elif "file kholo" in lower or "open file" in lower:
            fragment = lower.replace("file kholo", "").replace("open file", "").strip()
            matches = self.files.search(fragment, max_results=1)
            tool_output = self.files.open_file(matches[0]) if matches else "Ye file nahi mili."

        # System info
        elif "system info" in lower or "battery" in lower or "time kya hai" in lower or "abhi kya time hai" in lower:
            tool_output = self.sysinfo.status()

        # Clipboard
        elif "clipboard padho" in lower or "read clipboard" in lower:
            tool_output = self.clipboard.read()
        elif "clipboard mein copy" in lower or "copy to clipboard" in lower:
            text_to_copy = re.sub(r".*copy( karo)?", "", lower).strip() or user_input
            tool_output = self.clipboard.copy(text_to_copy)

        # Screen vision
        elif ("screen dekho" in lower or "analyze screen" in lower or "screen analyze" in lower) and self.tools.vision_manager:
            tool_output = self.tools.vision_manager.analyze_screen(user_input)
        elif ("screen dekho" in lower or "analyze screen" in lower or "screen analyze" in lower):
            tool_output = "Screen dekhne ke liye 'pyautogui' install karo: pip install pyautogui"

        # Shopping: "amazon par running shoes khojo", "flipkart par jeans dhundo"
        elif "amazon" in lower or "flipkart" in lower:
            platform = "amazon" if "amazon" in lower else "flipkart"
            query = lower
            for word in [platform, "par", "pe", "khojo", "dhundo", "search", "kholo", "find"]:
                query = query.replace(word, " ")
            tool_output = self.tools.shop_search(platform, query.strip())

        # WhatsApp: "whatsapp 9876543210 kal milte hain 5 baje"
        elif "whatsapp" in lower:
            match = re.search(r"(\+?\d{10,15})\s+(.*)", user_input.replace("whatsapp", "", 1).strip())
            if match:
                number, message = match.group(1), match.group(2)
                tool_output = self.whatsapp.send_message(number, message)
            else:
                tool_output = "WhatsApp bhejne ke liye number aur message dono boliye — jaise 'whatsapp 9876543210 kal milte hain'"

        # Command Processing: Open / Play / Close apps
        elif "open" in lower or "kholo" in lower or "chalu karo" in lower:
            app_target = lower.replace("open", "").replace("kholo", "").replace("chalu karo", "").strip()
            tool_output = self.tools.execute_app_command("open", app_target)
        elif "play" in lower or "gana" in lower or "song" in lower or "chalao" in lower:
            song_target = lower.replace("play", "").replace("gana", "").replace("song", "").replace("chalao", "").strip()
            # If they mentioned YouTube, actually play the video there
            # (via pywhatkit) instead of just searching on Spotify.
            if "youtube" in lower or "yt" in lower.split():
                song_target = song_target.replace("youtube", "").replace(" pe", "").replace(" par", "").strip()
                tool_output = self.tools.execute_app_command("play_youtube", song_target)
            else:
                tool_output = self.tools.execute_app_command("play", song_target)
        elif "close" in lower or "band karo" in lower:
            app_target = lower.replace("close", "").replace("band karo", "").strip()
            tool_output = self.tools.execute_app_command("close", app_target)

        # Research & Web Integration
        if any(w in lower for w in ["research", "search", "information", "news", "mausam", "kya chal raha hai", "batao"]):
            search_res = self.web.search_web(user_input)
            if search_res["status"] == "success":
                web_context = search_res["context"]

        # Save to Desktop or Excel Request Handling
        if "excel" in lower or "sheet" in lower or "spreadsheet" in lower:
            # Timestamped filename so each report is saved separately
            # instead of silently overwriting the previous one.
            report_name = f"Annu_Research_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}"
            tool_output += "\n" + self.workspace.create_professional_excel(
                report_name,
                [["Topic / Query", user_input], ["Timestamp", str(datetime.now())], ["Status", "Completed Successfully"]], 
                ["Field Name", "Details / Data"]
            )
        elif "save" in lower or "desktop par save" in lower:
            tool_output += "\n" + self.workspace.save_to_desktop("Annu_Note", f"Query: {user_input}\nTime: {datetime.now()}", "txt")

        self.memory.process_and_store(user_input)
        self.semantic_memory.add_memory(user_input, role="user")
        recalled_context = self.semantic_memory.recall(user_input)

        augmented_prompt = (
            f"User Input: '{user_input}'\n"
            f"Relevant Past Context: {recalled_context}\n"
            f"Web/Research Context: {web_context}\n"
            f"Tool Execution Output: {tool_output}\n\n"
            f"[System Instruction: Act as Annu-chan, an expert AI assistant. Provide a helpful, accurate, professional response in Hindi.]"
        )

        max_retries = 4
        for attempt in range(max_retries):
            try:
                response = self.chat_session.send_message(augmented_prompt)
                self.voice.speak(response.text)
                break
            except Exception as e:
                print(f"\n[DEBUG API ERROR]: {e}")
                error_str = str(e)
                is_quota = "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower()
                # 503/UNAVAILABLE = Google's servers are temporarily
                # overloaded ("high demand"), not a problem with our code —
                # this needs a retry-with-backoff too, not just quota errors.
                is_overloaded = (
                    "503" in error_str or "UNAVAILABLE" in error_str
                    or "overloaded" in error_str.lower() or "high demand" in error_str.lower()
                )

                if is_quota and self.model_manager.rotate_key():
                    self.chat_session = self.model_manager.new_chat()
                    continue

                if (is_quota or is_overloaded) and attempt < max_retries - 1:
                    delay = 3 * (2 ** attempt)  # 3s, 6s, 12s
                    self.logger.warning(f"Gemini temporarily unavailable, retrying in {delay}s (attempt {attempt + 1}/{max_retries})")
                    self.safe_update_display(f"Google server busy hai, {delay}s mein phir try kar rahi hoon...", "System")
                    time.sleep(delay)
                    continue

                if is_overloaded:
                    self.voice.speak("Google ke servers abhi busy hain. Thodi der baad phir try karo.")
                else:
                    self.voice.speak("Mujhe processing mein error aa rahi hai.")
                break

    def start_background_threads(self):
        self.running = True
        # Voice loop thread
        threading.Thread(target=self.background_voice_loop, daemon=True).start()
        # Morning 10 AM Scheduler Thread
        threading.Thread(target=self.background_scheduler_loop, daemon=True).start()
        # Reminder/task due-checker thread
        threading.Thread(target=self.background_reminder_loop, daemon=True).start()

    def background_reminder_loop(self):
        """Polls for due reminders every 20s and speaks them aloud."""
        while self.running:
            try:
                for task_id, description in self.tasks.due_tasks():
                    self.voice.speak(f"Reminder: {description}")
            except Exception as e:
                self.logger.error(f"Reminder loop error: {e}")
            time.sleep(20)

    def background_voice_loop(self):
        self.voice.speak("All core modules and research tools are online.")
        while self.running:
            try:
                user_input = self.voice.listen()
                if not user_input:
                    continue
                self.safe_update_display(user_input, "You")
                self.handle_user_input(user_input)
            except Exception as e:
                self.logger.error(f"Voice loop error: {e}")

    def background_scheduler_loop(self):
        """Runs automated tasks daily at 10:00 AM"""
        def morning_task():
            self.voice.speak("Good morning! Main aapki daily briefing research karke Excel mein taiyar kar rahi hoon.")

            rows = [
                ["Report Date", datetime.now().strftime("%A, %d %B %Y")],
                ["Generated At", datetime.now().strftime("%I:%M %p")],
            ]

            news_items = self.web.search_structured("India top news headlines today", max_results=5)
            if news_items:
                for idx, item in enumerate(news_items, 1):
                    title = (item.get("title") or "").strip()
                    body = (item.get("body") or "").strip()
                    summary = f"{title} — {body}" if title and body else (title or body)
                    rows.append([f"News {idx}", summary[:500]])
            else:
                rows.append(["News", "Abhi research fetch nahi ho payi — internet connection check karo."])

            pending_tasks = self.tasks.list_tasks()
            if pending_tasks:
                for task_id, description, remind_at, _done in pending_tasks[:5]:
                    when = f" (reminder: {remind_at})" if remind_at else ""
                    rows.append([f"Pending Task #{task_id}", f"{description}{when}"])
            else:
                rows.append(["Pending Tasks", "Koi pending task nahi hai."])

            result = self.workspace.create_professional_excel(
                f"Morning_Briefing_{datetime.now().strftime('%Y-%m-%d')}",
                rows,
                ["Section", "Details"],
            )
            self.safe_update_display(result, "System")

        schedule.every().day.at("10:00").do(morning_task)

        while self.running:
            schedule.run_pending()
            time.sleep(30)

def main():
    root = tk.Tk()
    app = AnnuAIBrainGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()